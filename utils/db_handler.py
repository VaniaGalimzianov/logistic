import io
from sqlalchemy import create_engine, Column, String, Integer, Float
from sqlalchemy.orm import sessionmaker, declarative_base
from typing import List, Dict, Any
import hashlib
from openpyxl import load_workbook
# from config import db_host, db_user, db_password, db_name


db_host = "127.0.0.1"
db_user = "postgres"
db_password = "############"
db_name = "workers"

# Создаем базовый класс модели
Base = declarative_base()

# Модель таблицы workers
class Worker(Base):
    __tablename__ = 'workers'

    worker_code = Column(Integer, primary_key=True)
    passwd = Column(String)
    name = Column(String)

# Модель таблицы cars
class Car(Base):
    __tablename__ = 'cars'

    id = Column(Integer, primary_key=True, autoincrement=True,)
    name = Column(String)
    payload = Column(Integer)
    length = Column(Float)
    width = Column(Float)
    height = Column(Float)
    status = Column(String, default="Свободно")

# Класс для работы с базой данных
class DbHandler:
    def __init__(self):
        db_url = f"postgresql://{db_user}:{db_password}@{db_host}/{db_name}"
        self.engine = create_engine(db_url)
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.session = Session()

    def check_credentials(self, worker_code, passwd) -> str | None:
        hashed_object = hashlib.md5(passwd.encode())
        passwd = hashed_object.hexdigest()

        worker = self.session.query(Worker).filter_by(
            worker_code=worker_code,
            passwd=passwd
        ).first()
    
        if worker:
            return worker.name
    

    def insert_car(self, name, payload, length, width, height) -> None:
        try:
            new_car = Car(
                name=name,
                payload=payload,
                length=length,
                width=width,
                height=height
            )
            
            self.session.add(new_car)
            self.session.commit()

            return 201

        except Exception as e:
            self.session.rollback()
            print(f"Ошибка при добавлении автомобиля: {e}")
            return 500
        
    
    def get_all_cars(self) -> List[Dict[str, Car]]:
        cars_data = self.session.query(Car).all()
        cars_list = []
        for car in cars_data:
            car_info = {
                'name': car.name,
                'payload': car.payload,
                'length': car.length,
                'width': car.width,
                'height': car.height,
                'status': car.status
            }
            cars_list.append(car_info)
        
        return cars_list

    def file_parser(self, file) -> List[Car]:
        cars_list = []
        wb = load_workbook(file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # Проверяем, что хотя бы одна ячейка в строке не пустая
                car = Car(
                    name=row[headers.index('Авто')],
                    payload=row[headers.index('Грузоподъемность')],
                    length=row[headers.index('Длина')],
                    width=row[headers.index('Ширина')],
                    height=row[headers.index('Высота')]
                )
                cars_list.append(car)
        
        return cars_list[:-1]

    

    def add_car_file(self, file) -> None:
        try:
            car_list = self.file_parser(io.BytesIO(file))

            for car_info in car_list:
                new_car = Car(
                    name=car_info.name,
                    payload=car_info.payload,
                    length=car_info.length,
                    width=car_info.width,
                    height=car_info.height
                )
                self.session.add(new_car)
            self.session.commit()

            return 201
        
        except Exception as e:
            self.session.rollback()
            print(f"Ошибка при добавлении автомобиля: {e}")
            return 500

    

    def find_optimal_cars(self,
                          payload_needed,
                          length_needed,
                          width_needed,
                          height_needed
                        ) -> List[Dict[str, Any]]:
        
        optimal_cars = []

        all_cars = self.session.query(Car).filter_by(status="Свободно").all()

        for car in all_cars:
            car_info = {
                'name': car.name,
                'payload': car.payload,
                'length': car.length,
                'width': car.width,
                'height': car.height,
            }
            if car.payload >= payload_needed:
                payload_difference = abs(car.payload - payload_needed)
                length_difference = abs(car.length - length_needed)
                width_difference = abs(car.width - width_needed)
                height_difference = abs(car.height - height_needed)

                total_difference = payload_difference + length_difference + width_difference + height_difference

                optimal_cars.append((car_info, total_difference))

        optimal_cars.sort(key=lambda x: x[1])  # Сортируем по возрастанию различий

        return [car[0] for car in optimal_cars[:3]]
    

    def update_car_status(self, name, payload, length, width, height, status) -> None:
        car = self.session.query(Car).filter_by(
            name=name,
            payload=payload,
            length=length,
            width=width,
            height=height
        ).first()

        if car:
            car.status = 'Занято'
            self.session.commit()
            print(f"Статус машины {name} изменен на 'Занято'")
        else:
            print("Машина не найдена в базе данных")
        
        self.session.close()
